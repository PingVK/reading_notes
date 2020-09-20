"""
    1         2      3     4      5         6       7      8     9      10      11     12        13     14   15
图片路径  图片名字  姓名  性别  出生日期  证件类型  证件号  民族  籍贯  家庭住址  编号  人员特征  人员类型  备注  id
"""
from openpyxl import load_workbook
import xlsxwriter
from fake_data import *


def get_info_list():
    name = get_name(length=3)  # 姓名
    sex = get_sex()  # 性别
    birthday = get_birthday(fmt='%Y/%m/%d')  # 出生日期
    id_type = get_id_type()  # 证件类型
    id_number = get_id_number()  # 证件号
    ethnicity = get_ethnicity()  # 民族
    birth_place = get_address(city_flag=False, street_flag=False, street_num_flag=False)  # 籍贯
    address = get_address()  # 家庭住址
    serial_number = get_serial_number()  # 编号
    member_characteristic = get_member_characteristic()  # 人员特征
    member_type = get_member_type()  # 人员类型
    remark = get_remark()  # 备注
    return (name, sex, birthday, id_type, id_number, ethnicity, birth_place, address, serial_number,
            member_characteristic, member_type, remark)


def modify_excel(open_path, save_path):
    print('loading...')
    # 读取EXCEL表
    wb = load_workbook(open_path, read_only=True)
    ws = wb['Sheet1']
    ws_iter = ws.rows

    # 新建EXCEL表
    workbook = xlsxwriter.Workbook(save_path)
    sheet = workbook.add_worksheet()

    # 写入表头
    row_head = next(ws_iter)
    for i in range(15):
        sheet.write(0, i, row_head[i].value)
    # 开始逐行写入
    row = 1
    print('writing...')
    for row_info in ws_iter:
        info = get_info_list()
        for col in (0, 1, 14):
            sheet.write(row, col, row_info[col].value)
        for col in range(2, 14):
            sheet.write(row, col, info[col-2])
        row += 1
        if row % 10000 == 0:
            print('%d行数据已写入' % row)
    print('saving...')
    workbook.close()
    print('Success')


if __name__ == '__main__':
    modify_excel(r'D:\qa-picture\1W\batch_file.xlsx', r'D:\qa-picture\1W\batch_file.xlsx')
